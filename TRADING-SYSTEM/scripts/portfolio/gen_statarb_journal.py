import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"D:\Vault\Project\R_projekt\R-trade\R_trade_proj_sin-arb\TRADING-SYSTEM\05-EXCEL\stat-arb-journal.xlsx"
os.makedirs(os.path.dirname(OUT), exist_ok=True)

AR = "Arial"
BLUE = Font(name=AR, size=10, color="0000FF")        # ввод
BLK  = Font(name=AR, size=10, color="000000")        # формула
BOLD = Font(name=AR, size=10, bold=True)
TITLE= Font(name=AR, size=14, bold=True, color="FFFFFF")
HEADW= Font(name=AR, size=10, bold=True, color="FFFFFF")
HINT = Font(name=AR, size=9, italic=True, color="595959")
HEADFILL = PatternFill("solid", fgColor="1F4E78")
SUBFILL  = PatternFill("solid", fgColor="DDEBF7")
INFILL   = PatternFill("solid", fgColor="FFF2CC")    # лёгкий жёлтый под ввод
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin,right=thin,top=thin,bottom=thin)

def cell(ws,ref,val=None,font=BLK,fmt=None,fill=None,align=None,border=False):
    c=ws[ref]
    if val is not None: c.value=val
    c.font=font
    if fmt: c.number_format=fmt
    if fill: c.fill=fill
    if align: c.alignment=Alignment(horizontal=align,vertical="center")
    if border: c.border=BORD
    return c

PX='#,##0.00'; USD='$#,##0.00'; USD0='$#,##0'; R4='0.0000'; Z2='0.00'; PCT='0.0%'

wb=Workbook()

# ════════════════ ЛИСТ 1: КАЛЬКУЛЯТОР ════════════════
ws=wb.active; ws.title="Калькулятор"
ws.sheet_view.showGridLines=False
for col,w in {"A":30,"B":4,"C":26,"D":26,"E":3}.items(): ws.column_dimensions[col].width=w
ws.merge_cells("A1:D1"); cell(ws,"A1","СТАТ-АРБ — КАЛЬКУЛЯТОР (кросс-венью спред)",TITLE,fill=HEADFILL,align="left")
ws.row_dimensions[1].height=24
cell(ws,"A2","Жёлтое = ввод (синий шрифт). Чёрное = формула. Зеркалит вкладку «Стат-арб 🔱».",HINT)

# ноги
cell(ws,"A3","ПАРАМЕТР",HEADW,fill=HEADFILL,border=True); cell(ws,"C3","Нога A",HEADW,fill=HEADFILL,align="center",border=True); cell(ws,"D3","Нога B",HEADW,fill=HEADFILL,align="center",border=True)
rows=[("Метка / тикер","XAUTUSDT.bybit","XAUEUR.infra",None,True),
      ("Площадка","Bybit","Infra Capital",None,True),
      ("Цена","4021.2","3724",PX,True),
      ("FX → USD","1","1.08",R4,True),
      ("$ / пункт","1","1",R4,True),
      ("Вола %","15","15","0.0",True)]
r=4
for lab,va,vb,fmt,isin in rows:
    cell(ws,f"A{r}",lab,BOLD,border=True)
    for col,v in (("C",va),("D",vb)):
        try: v=float(v)
        except: pass
        cell(ws,f"{col}{r}",v,BLUE if isin else BLK,fmt,INFILL if isin else None,border=True)
    r+=1
cell(ws,"A10","TV-символ",BOLD,border=True)
cell(ws,"C10","OANDA:XAUUSD",BLUE,fill=INFILL,border=True); cell(ws,"D10","OANDA:XAUEUR*FX_IDC:EURUSD",BLUE,fill=INFILL,border=True)
cell(ws,"A11","Цена в USD",BOLD,border=True); cell(ws,"C11","=C6*C7",BLK,PX,border=True); cell(ws,"D11","=D6*D7",BLK,PX,border=True)
cell(ws,"A12","Ноционал / ед, $",BOLD,border=True); cell(ws,"C12","=C11*C8",BLK,USD,border=True); cell(ws,"D12","=D11*D8",BLK,USD,border=True)

# параметры
cell(ws,"A14","ПАРАМЕТРЫ",HEADW,fill=HEADFILL,border=True); cell(ws,"C14","",fill=HEADFILL,border=True); cell(ws,"D14","",fill=HEADFILL,border=True)
P=[("Метод ратио","notional",None),("β ручной (ед B на ед A)",1,R4),("Капитал / сторона, $",1000,USD0)]
r=15
for lab,v,fmt in P:
    cell(ws,f"A{r}",lab,BOLD,border=True); cell(ws,f"C{r}",v,BLUE,fmt,INFILL,border=True); cell(ws,f"D{r}","",border=True)
    r+=1
cell(ws,"A18","Хедж-ратио h (ед B на 1 A)",BOLD,border=True)
cell(ws,"C18",'=IF(C15="notional",C12/D12,IF(C15="vol",(C12*C9)/(D12*D9),C16))',BLK,R4,border=True); cell(ws,"D18","",border=True)
cell(ws,"A19","Units A",BOLD,border=True); cell(ws,"C19","=C17/C12",BLK,R4,border=True); cell(ws,"D19","",border=True)
cell(ws,"A20","Units B",BOLD,border=True); cell(ws,"C20","=C19*C18",BLK,R4,border=True); cell(ws,"D20","",border=True)
cell(ws,"A21","Ноционал A, $",BOLD,border=True); cell(ws,"C21","=C19*C12",BLK,USD0,border=True); cell(ws,"D21","",border=True)
cell(ws,"A22","Ноционал B, $",BOLD,border=True); cell(ws,"C22","=C20*D12",BLK,USD0,border=True); cell(ws,"D22","",border=True)

# сигнал
cell(ws,"A24","СИГНАЛ",HEADW,fill=HEADFILL,border=True); cell(ws,"C24","",fill=HEADFILL,border=True); cell(ws,"D24","",fill=HEADFILL,border=True)
cell(ws,"A25","Режим (diff/ratio/pct)",BOLD,border=True); cell(ws,"C25","diff",BLUE,fill=INFILL,border=True); cell(ws,"D25","",border=True)
cell(ws,"A26","Спред авто",BOLD,border=True); cell(ws,"C26",'=IF(C25="ratio",C11/D11,IF(C25="pct",(C11/D11-1)*100,C11-C18*D11))',BLK,PX,border=True); cell(ws,"D26","",border=True)
cell(ws,"A27","Спред сейчас (пусто=авто)",BOLD,border=True); cell(ws,"C27","",BLUE,PX,INFILL,border=True); cell(ws,"D27","",border=True)
cell(ws,"A28","Спред используемый",BOLD,border=True); cell(ws,"C28",'=IF(C27="",C26,C27)',BLK,PX,border=True); cell(ws,"D28","",border=True)
cell(ws,"A29","μ среднее",BOLD,border=True); cell(ws,"C29",0,BLUE,PX,INFILL,border=True); cell(ws,"D29","",border=True)
cell(ws,"A30","σ спреда",BOLD,border=True); cell(ws,"C30",5,BLUE,PX,INFILL,border=True); cell(ws,"D30","",border=True)
cell(ws,"A31","Порог Z",BOLD,border=True); cell(ws,"C31",2,BLUE,Z2,INFILL,border=True); cell(ws,"D31","",border=True)
cell(ws,"A32","Z-score",BOLD,border=True); cell(ws,"C32","=IF(C30=0,0,(C28-C29)/C30)",BLK,Z2,border=True); cell(ws,"D32","",border=True)
cell(ws,"A33","СИГНАЛ",BOLD,border=True)
cell(ws,"C33",'=IF(C32>=C31,"ШОРТ A / ЛОНГ B (спред дорог)",IF(C32<=-C31,"ЛОНГ A / ШОРТ B (спред дёшев)","ждать (внутри полосы)"))',Font(name=AR,size=10,bold=True),border=True); ws.merge_cells("C33:D33")
cell(ws,"A34","P&L цель к μ, $",BOLD,border=True); cell(ws,"C34",'=IF(C25="diff",ABS(C28-C29)*C19*C8,"оцени по графику")',BLK,USD,border=True); cell(ws,"D34","",border=True)

cell(ws,"A36","TV-ФОРМУЛА",HEADW,fill=HEADFILL,border=True); cell(ws,"C36","",fill=HEADFILL,border=True); cell(ws,"D36","",fill=HEADFILL,border=True)
cell(ws,"A37","Формула спреда",BOLD,border=True)
cell(ws,"C37",'=IF(C25="ratio",C10&"/("&D10&")",IF(C25="pct","("&C10&"/("&D10&")-1)*100",C10&" - "&TEXT(C18,"0.0000")&"*("&D10&")"))',BLK,border=True); ws.merge_cells("C37:D37")
cell(ws,"A38","→ вставь как символ в TradingView; снимай μ/σ с этого графика",HINT)

dv1=DataValidation(type="list",formula1='"notional,vol,beta"',allow_blank=False); ws.add_data_validation(dv1); dv1.add(ws["C15"])
dv2=DataValidation(type="list",formula1='"diff,ratio,pct"',allow_blank=False); ws.add_data_validation(dv2); dv2.add(ws["C25"])

# ════════════════ ЛИСТ 2: ЖУРНАЛ ════════════════
jw=wb.create_sheet("Журнал"); jw.sheet_view.showGridLines=False
heads=["Дата","Пара","Класс актива","Нога A","Площ. A","Нога B","Площ. B","Метод","Направление",
       "Хедж-ратио","Спред вход","μ","σ","Z вход","Units A","Units B","Ноционал/стор $",
       "Время вход","Спред выход","Время выход","Реализ.P&L $","Фи Bybit $","Своп/фи CFD $","Нетто P&L $","Статус","Заметки"]
widths=[11,16,13,16,12,16,12,10,22,11,11,8,8,8,9,9,14,11,11,11,13,11,13,13,9,30]
for i,(h,w) in enumerate(zip(heads,widths),1):
    c=jw.cell(row=1,column=i,value=h); c.font=HEADW; c.fill=HEADFILL; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=BORD
    jw.column_dimensions[chr(64+i) if i<=26 else "A"+chr(64+i-26)].width=w
jw.row_dimensions[1].height=30
jw.freeze_panes="A2"

# пример строки
ex=["2026-06-30","XAUT/XAUEUR","Золото","XAUTUSDT.bybit","Bybit","XAUEUR.infra","Infra","notional",
    "ШОРТ A / ЛОНГ B",1.0,15,0,5,None,0.2487,0.2486,1000,"08:20","", "",None,0.20,1.50,None,None,"тест-сетап"]
NROWS=60
for rr in range(2,2+NROWS):
    for ci in range(1,27):
        c=jw.cell(row=rr,column=ci); c.border=BORD; c.font=BLK
    # пример только в первой строке
    if rr==2:
        for ci,v in enumerate(ex,1):
            if v is not None: jw.cell(row=rr,column=ci,value=v)
    # синий шрифт для столбцов ввода
    for ci in [1,2,3,4,5,6,7,8,9,10,11,12,13,15,16,17,18,19,20,21,22,23,26]:
        jw.cell(row=rr,column=ci).font=BLUE
    # формулы (чёрные): N=Z вход, X=нетто, Y=статус
    jw.cell(row=rr,column=14,value=f'=IF(OR(K{rr}="",M{rr}="",M{rr}=0),"",(K{rr}-L{rr})/M{rr})').font=BLK   # Z вход
    jw.cell(row=rr,column=24,value=f'=IF(U{rr}="","",U{rr}-IF(V{rr}="",0,V{rr})-IF(W{rr}="",0,W{rr}))').font=BLK  # нетто
    jw.cell(row=rr,column=25,value=f'=IF(A{rr}="","",IF(S{rr}="","открыт","закрыт"))').font=BLK   # статус
    jw.cell(row=rr,column=14).number_format=Z2
    for ci in [10]: jw.cell(row=rr,column=ci).number_format=R4
    for ci in [11,12,13,19]: jw.cell(row=rr,column=ci).number_format=PX
    for ci in [15,16]: jw.cell(row=rr,column=ci).number_format=R4
    for ci in [17,21,22,23,24]: jw.cell(row=rr,column=ci).number_format=USD0

# ════════════════ ЛИСТ 3: СВОДКА ════════════════
sw=wb.create_sheet("Сводка"); sw.sheet_view.showGridLines=False
sw.column_dimensions["A"].width=30; sw.column_dimensions["B"].width=16
sw.merge_cells("A1:B1"); cell(sw,"A1","СВОДКА",TITLE,fill=HEADFILL,align="left"); sw.row_dimensions[1].height=24
J="Журнал!"
stats=[("Всего сетапов",f'=COUNTA({J}A2:A61)',"0"),
       ("Открыто",f'=COUNTIF({J}Y2:Y61,"открыт")',"0"),
       ("Закрыто",f'=COUNTIF({J}Y2:Y61,"закрыт")',"0"),
       ("Выигрышных",f'=COUNTIFS({J}Y2:Y61,"закрыт",{J}X2:X61,">0")',"0"),
       ("Win rate",'=IF(B5=0,0,B6/B5)',PCT),
       ("Нетто P&L, $",f'=SUM({J}X2:X61)',USD),
       ("Валовой P&L, $",f'=SUM({J}U2:U61)',USD),
       ("Сумма фи (Bybit+CFD), $",f'=SUM({J}V2:V61)+SUM({J}W2:W61)',USD),
       ("Средний Z входа (закрытых)",f'=IFERROR(AVERAGEIFS({J}N2:N61,{J}Y2:Y61,"закрыт"),0)',Z2),
       ("Лучший сетап, $",f'=IFERROR(MAX({J}X2:X61),0)',USD),
       ("Худший сетап, $",f'=IFERROR(MIN({J}X2:X61),0)',USD)]
r=3
for lab,fml,fmt in stats:
    cell(sw,f"A{r}",lab,BOLD,fill=SUBFILL,border=True)
    cell(sw,f"B{r}",fml,BLK,fmt if fmt not in("0",) else "0",border=True)
    r+=1
cell(sw,f"A{r+1}","P&L по классу актива (заполни классы в Журнале, столбец C):",HINT)
r+=2
for cls in ["Золото","Серебро","Платина","BTC","ETH","Валюта"]:
    cell(sw,f"A{r}",cls,BLK,border=True)
    cell(sw,f"B{r}",f'=SUMIF({J}C2:C61,A{r},{J}X2:X61)',BLK,USD,border=True)
    r+=1

# ════════════════ ЛИСТ 4: ЗАДАЧИ ════════════════
qw=wb.create_sheet("Задачи"); qw.sheet_view.showGridLines=False
for col,w in {"A":4,"B":12,"C":80,"D":13,"E":5,"F":48,"G":15}.items(): qw.column_dimensions[col].width=w
qw.merge_cells("A1:F1"); cell(qw,"A1","ЗАДАЧИ — триангулярный стат-арб (детсад → практик)",TITLE,fill=HEADFILL,align="left"); qw.row_dimensions[1].height=24
cell(qw,"A2","Вписывай жёлтый «Твой ответ» → ✓ загорится само (допуск ±1%). Правильный ответ скрыт в столбце G — разверни, только если застрял.",HINT)
for i,htxt in enumerate(["№","Уровень","Вопрос","Твой ответ","✓","Подсказка / формула","Ответ"],1):
    c=qw.cell(row=3,column=i,value=htxt); c.font=HEADW; c.fill=HEADFILL; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=BORD
qw.row_dimensions[3].height=22; qw.freeze_panes="A4"

P=[
("🧸 Детсад","1 EUR = 1.08 USD. Сколько USD за 50 EUR?",True,"1.08*50","умножь курс на количество"),
("🧸 Детсад","1 USD = 150 JPY. Сколько JPY за 100 USD?",True,"150*100","умножь"),
("🧸 Детсад","1 EUR = 1.08 USD и 1 USD = 150 JPY. Сколько JPY за 1 EUR через доллар?",True,"1.08*150","перемножь два курса = синтетический EURJPY"),
("🧸 Детсад","У тебя 16 200 JPY, курс EURJPY = 162. Сколько это EUR?",True,"16200/162","дели сумму на курс"),
("🧸 Детсад","Старт 100 EUR → доллары (×1.08) → йены (×150) → назад в EUR по EURJPY=162.30. Сколько EUR в конце? (<100 = потерял)",True,"100*1.08*150/162.30","умножай при продаже, дели при выкупе"),
("🎒 Школа","Синтетика EURJPY = 1.08×150. Прямая котировка = 162.30. Гэп в пунктах (йенах)?",True,"162.30-1.08*150","прямая − синтетика"),
("🎒 Школа","Тот же гэп в процентах от синтетики?",True,"(162.30/(1.08*150)-1)*100","(прямая/синт − 1)×100"),
("🎒 Школа","EURJPY с 2 знаками, 1 пункт = 0.01. Гэп 0.30 = сколько пунктов?",True,"0.30/0.01","дели гэп на размер пункта"),
("🎒 Школа","Круг проходит 3 ноги, каждая 0.04% туда-обратно. Сколько % съест весь круг?",True,"3*0.04","сложи издержки трёх ног"),
("🎒 Школа","Край круга 0.185% против издержек 0.12%. Сколько нетто остаётся, %?",True,"0.185-0.12","край − издержки; >0 = живой"),
("🎒 Школа","Лог-остаток x = ln(162.30) − ln(1.08) − ln(150). Посчитай.",True,"LN(162.30)-LN(1.08)-LN(150)","натуральный лог, знаки (+,−,−)"),
("🎓 Универ","Коинтегрирующий вектор на (ln EURJPY, ln EURUSD, ln USDJPY)? Три числа.",False,"(1; -1; -1)","из тождества c=e·j → ln c − ln e − ln j = 0"),
("🎓 Универ","Лонг 100 000 EUR ноционала в EURJPY. Какой USD-ноционал в EURUSD-ноге обнуляет EUR? (×1.08)",True,"100000*1.08","EUR-ноционал × курс EURUSD"),
("🎓 Универ","Та же корзина: JPY-ноционал в USDJPY-ноге? (EUR × 1.08 × 150)",True,"100000*1.08*150","прогон по цепочке валют 1 : e : e·j"),
("🎓 Универ","Остаток: μ=0, σ=0.0003 (лог), текущий x=0.0008. z-score?",True,"(0.0008-0)/0.0003","(x − μ) / σ"),
("🎓 Универ","Порог входа z=2, текущий z=2.667 — сигнал есть? (1=да, 0=нет)",True,"IF(2.667>=2,1,0)","сравни |z| с порогом"),
("🎓 Универ","OU-процесс, скорость возврата κ=0.5 в минуту. Полупериод возврата (мин)?",True,"LN(2)/0.5","half-life = ln2 / κ"),
("🎓 Универ","Лог-остаток обратно в %: x=0.001849 → (e^x − 1)×100 = ?",True,"(EXP(0.001849)-1)*100","exp(x) − 1, ×100"),
("🎓 Универ","Почему σ остатка (0.0003) ≪ σ ноги (~0.01)? Словами.",False,"ноги коинтегрированы (ходят вместе) → дисперсии в комбинации (1,−1,−1) почти гасятся",": Var(a−b−c) при корреляции →1 стремится к 0"),
("💼 Практик","Круг по стакану: продаёшь EUR по EURUSD bid 1.0800 → USD по USDJPY bid 150.00 → выкуп EUR по EURJPY ask 162.40. Сколько EUR на 1 EUR?",True,"1.0800*150.00/162.40","перемножь исполняемые курсы; выкуп = деление на ask"),
("💼 Практик","Тот круг — сколько потерял, %?",True,"(1.0800*150.00/162.40-1)*100","(результат − 1)×100; минус = убыток (издержки съели)"),
("💼 Практик","Кросс-венью: EURJPY на площадке A = 162.30, на B = 162.10. Базис в пунктах?",True,"162.30-162.10","разница цены ОДНОЙ пары на двух площадках — вот где край живёт"),
("💼 Практик","CIP-форвард EURUSD на год: спот 1.0800, r_USD=5%, r_EUR=3%. Форвард = спот×(1+r_USD)/(1+r_EUR).",True,"1.0800*(1+0.05)/(1+0.03)","паритет покрытых процентных ставок"),
("💼 Практик","Рыночный форвард = 1.0980. Кросс-валютный базис = рынок − CIP-форвард. Посчитай.",True,"1.0980-1.0800*(1+0.05)/(1+0.03)","market − CIP = xccy basis, реальный край банковских десков"),
("💼 Практик","Цикл 4 курсов: EUR→USD 1.08, USD→JPY 150, JPY→GBP 0.0053, GBP→EUR 1.16. Произведение по кругу?",True,"1.08*150*0.0053*1.16","произведение по замкнутому циклу; ≠1 = арбитраж (>1 прибыль). Это Bellman–Ford"),
("💼 Практик","Годовой Шарп: край 2 б.п./цикл, σ 3 б.п., 50 циклов/день, 252 дня. Sharpe = (2/3)×√(50×252).",True,"(2/3)*SQRT(50*252)","Sharpe = средн/σ × √(сделок в год). Заоблачный → оттого за край дерётся HFT; в ликвиде он ≈0 после издержек"),
]
r=4
for lvl,q,isnum,ans,hnt in P:
    cell(qw,f"A{r}",r-3,BLK,align="center",border=True)
    cell(qw,f"B{r}",lvl,BLK,border=True)
    c=cell(qw,f"C{r}",q,BLK,border=True); c.alignment=Alignment(wrap_text=True,vertical="top")
    cell(qw,f"D{r}",None,BLUE,fill=INFILL,border=True)
    fc=cell(qw,f"F{r}",hnt,HINT,border=True); fc.alignment=Alignment(wrap_text=True,vertical="top")
    if isnum:
        cell(qw,f"G{r}","="+ans,BLK,border=True)
        cell(qw,f"E{r}",f'=IF(D{r}="","",IF(ABS(D{r}-G{r})<=ABS(G{r})*0.01+0.0001,"✓","✗"))',BLK,align="center",border=True)
    else:
        cell(qw,f"G{r}",ans,BLK,border=True)
        cell(qw,f"E{r}","устно",HINT,align="center",border=True)
    qw.row_dimensions[r].height=30
    r+=1
qw.column_dimensions["G"].hidden=True

wb.save(OUT)
print("SAVED",OUT)
