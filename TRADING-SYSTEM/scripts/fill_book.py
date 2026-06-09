"""
fill_book.py — заполняет TRADING_BOOK.xlsx реальными данными из Bybit API
Запуск: python TRADING-SYSTEM/scripts/fill_book.py

Требует:
  pip install openpyxl
  C:\\Users\\DoOs\\.bybit\\credentials.json  (api_key + api_secret, read-only)

Что заполняет:
  LEGS    — все открытые позиции (perp linear + spot + options)
  DAILY   — снапшот баланса на сегодня
  CONSTRUCTS — нетто-дельта и PnL по конструктам
"""

import sys, os, json, hmac, hashlib, time, urllib.request, urllib.parse
from datetime import date, datetime

sys.stdout.reconfigure(encoding="utf-8")

# ── пути ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
BOOK_PATH   = os.path.normpath(os.path.join(SCRIPT_DIR, "..", "..", "TRADING_BOOK.xlsx"))
CREDS_PATH  = r"C:\Users\DoOs\.bybit\credentials.json"
BASE_URL    = "https://api.bybit.com"

# ── цвета (тёмная тема, дублируем из gen_excel) ───────────────────────────────
C_BG    = "1E1E2E"; C_HEADER = "2A2A3E"; C_ACCENT = "F0B429"
C_GREEN = "3DDC97"; C_RED    = "FF6B6B"; C_BLUE   = "7EB8F7"
C_GRAY  = "3A3A4E"; C_TEXT   = "E0E0F0"

# ── Bybit API ─────────────────────────────────────────────────────────────────
def load_creds():
    if not os.path.exists(CREDS_PATH):
        print(f"❌  Файл ключей не найден: {CREDS_PATH}")
        sys.exit(1)
    with open(CREDS_PATH, encoding="utf-8") as f:
        d = json.load(f)
    key = d.get("api_key",""); secret = d.get("api_secret","")
    if not key or not key.isascii() or len(key) < 10:
        print("❌  Вставь реальные API-ключи в", CREDS_PATH)
        sys.exit(1)
    return key, secret

def sign(secret, ts, key, recv_window, params_str):
    msg = f"{ts}{key}{recv_window}{params_str}"
    return hmac.new(secret.encode("utf-8"), msg.encode("utf-8"), hashlib.sha256).hexdigest()

def bybit_get(path, params, key, secret):
    ts   = str(int(time.time() * 1000))
    recv = "20000"
    qs   = urllib.parse.urlencode(params)
    sig  = sign(secret, ts, key, recv, qs)
    headers = {
        "X-BAPI-API-KEY":     key,
        "X-BAPI-TIMESTAMP":   ts,
        "X-BAPI-SIGN":        sig,
        "X-BAPI-RECV-WINDOW": recv,
        "Content-Type":       "application/json",
    }
    url = f"{BASE_URL}{path}?{qs}"
    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8","replace")
        print(f"  HTTP {e.code} {path}: {body[:200]}")
        return None

def fetch_all(path, params, key, secret, max_pages=5):
    """Пагинация через cursor."""
    results = []
    cursor  = None
    for _ in range(max_pages):
        p = dict(params)
        if cursor:
            p["cursor"] = cursor
        data = bybit_get(path, p, key, secret)
        if not data or data.get("retCode") != 0:
            break
        lst = (data.get("result") or {}).get("list") or []
        results.extend(lst)
        cursor = (data.get("result") or {}).get("nextPageCursor")
        if not cursor:
            break
    return results

# ── классификация конструктов ─────────────────────────────────────────────────
def classify(sym, category, size_sign):
    """Возвращает (construct_id, leg_role) или (None, None) если не классифицировано."""
    sym = sym.upper()
    # C1: XAUTUSDT spot long + XAUTUSDT perp short
    if sym == "XAUTUSDT" and category == "spot":
        return "C1", "spot_long"
    if sym == "XAUTUSDT" and category == "linear" and size_sign < 0:
        return "C1", "perp_short"
    # C2: XAUUSDT perp (reversion) + XAGUSDT perp (trend)
    if sym == "XAUUSDT" and category == "linear":
        return "C2", "reversion_perp"
    if sym == "XAGUSDT" and category == "linear":
        return "C2", "trend_perp"
    # C3: XAUTUSDT perp long (ratio leg) or XAGUSDT short (ratio hedge)
    if sym == "XAUTUSDT" and category == "linear" and size_sign > 0:
        return "C3", "ratio_long"
    # C3 options
    if category == "option" and "XAUT" in sym:
        return "C3", "otm_bought" if size_sign > 0 else "otm_sold"
    # C4: не будет в позициях пока нет живых синт-фьючей — помечаем вручную
    return "?", "unclassified"

# ── Excel helpers ─────────────────────────────────────────────────────────────
def load_book():
    try:
        import openpyxl
        return openpyxl.load_workbook(BOOK_PATH)
    except Exception as e:
        print(f"❌  Не удалось открыть {BOOK_PATH}: {e}")
        sys.exit(1)

def style_row(ws, row, bg, values, fmts=None):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    fill = PatternFill("solid", fgColor=bg)
    font = Font(color=C_TEXT, size=10, name="Consolas")
    aln  = Alignment(horizontal="left", vertical="center")
    s    = Side(style="thin", color="444466")
    brd  = Border(left=s, right=s, top=s, bottom=s)
    for c, val in enumerate(values, 1):
        cell = ws.cell(row, c, value=val)
        cell.fill = fill; cell.font = font
        cell.alignment = aln; cell.border = brd
        if fmts and c-1 < len(fmts) and fmts[c-1]:
            cell.number_format = fmts[c-1]

def find_or_add_leg_row(ws, leg_id):
    """Ищет строку по leg_id (col B), возвращает номер строки."""
    for r in range(2, ws.max_row + 2):
        v = ws.cell(r, 2).value
        if v is None or v == leg_id:
            return r
    return ws.max_row + 1

# ── главная логика ────────────────────────────────────────────────────────────
def main():
    key, secret = load_creds()
    print("🔑  Ключи загружены, запрашиваем данные Bybit...")

    # 1. Баланс
    bal_data = bybit_get("/v5/account/wallet-balance",
                         {"accountType":"UNIFIED"}, key, secret)
    eq = im = mm = 0.0
    if bal_data and bal_data.get("retCode") == 0:
        acc = (bal_data["result"]["list"] or [{}])[0]
        eq  = float(acc.get("totalEquity",0) or 0)
        im  = float(acc.get("totalInitialMargin",0) or 0)
        mm  = float(acc.get("totalMaintenanceMargin",0) or 0)
        print(f"  Эквити: {eq:.2f} USDT  |  IM: {im:.2f}  |  MM: {mm:.2f}")
    else:
        print("  ⚠️  Баланс не получен — проверь ключи")

    # 2. Позиции linear (USDT)
    pos_linear = fetch_all("/v5/position/list",
                           {"category":"linear","settleCoin":"USDT","limit":"200"},
                           key, secret)
    pos_linear = [p for p in pos_linear if float(p.get("size",0) or 0) != 0]
    print(f"  Linear позиций: {len(pos_linear)}")

    # 3. Спот (через wallet coins)
    spot_coins = []
    if bal_data and bal_data.get("retCode") == 0:
        coins = (bal_data["result"]["list"] or [{}])[0].get("coin",[])
        for c in coins:
            qty = float(c.get("walletBalance",0) or 0)
            if qty > 0 and c["coin"] in ("XAUT","XAG","BTC","ETH","SOL"):
                spot_coins.append(c)
        print(f"  Спот монет: {len(spot_coins)}")

    # 4. Опционы
    pos_options = fetch_all("/v5/position/list",
                            {"category":"option","limit":"200"},
                            key, secret)
    pos_options = [p for p in pos_options if float(p.get("size",0) or 0) != 0]
    print(f"  Опционных позиций: {len(pos_options)}")

    # 5. Тикеры для текущих цен
    prices = {}
    for sym in ("XAUTUSDT","XAUUSDT","XAGUSDT"):
        td = bybit_get("/v5/market/tickers",
                       {"category":"linear","symbol":sym}, key, secret)
        if td and td.get("retCode") == 0:
            lst = td["result"]["list"]
            if lst:
                prices[sym] = float(lst[0].get("lastPrice",0) or 0)
    print(f"  Цены: { {k:round(v,3) for k,v in prices.items()} }")

    # ── Открываем книгу ───────────────────────────────────────────────────────
    wb = load_book()
    ws_legs   = wb["LEGS"]
    ws_daily  = wb["DAILY"]
    ws_con    = wb["CONSTRUCTS"]

    ROW_COLORS = [C_BG, "1A1A2E"]

    # ── Заполняем LEGS ────────────────────────────────────────────────────────
    # Очищаем старые данные (строки 2+)
    for r in range(2, ws_legs.max_row + 1):
        for c in range(1, 18):
            ws_legs.cell(r, c).value = None

    legs_rows = []

    # linear perp
    for p in pos_linear:
        sym  = p["symbol"]
        side = p["side"]          # Buy / Sell
        size = float(p["size"] or 0)
        if side == "Sell": size = -size
        ep   = float(p.get("avgPrice",0) or 0)
        cp   = prices.get(sym, float(p.get("markPrice",0) or 0))
        uPnl = float(p.get("unrealisedPnl",0) or 0)
        delta_raw = float(p.get("delta", 1 if side=="Buy" else -1) or (1 if side=="Buy" else -1))
        cid, role = classify(sym, "linear", 1 if side=="Buy" else -1)
        leg_id = f"{cid}_{role[:8].upper()}"
        notional = abs(size) * ep
        legs_rows.append([
            cid, leg_id, role, sym, "perp",
            "long" if side=="Buy" else "short",
            abs(size), ep, cp, round(notional,2),
            round(uPnl,4), delta_raw, 0, 0, 0, "open", ""
        ])

    # spot coins
    for c in spot_coins:
        coin = c["coin"]
        sym  = coin + "USDT"
        qty  = float(c.get("walletBalance",0) or 0)
        ep   = float(c.get("avgPrice", prices.get(sym, 0)) or 0)
        cp   = prices.get(sym, ep)
        uPnl = (cp - ep) * qty if ep else 0
        cid, role = classify(sym, "spot", 1)
        leg_id = f"{cid}_{role[:8].upper()}_SPOT"
        legs_rows.append([
            cid, leg_id, role, f"{coin} (spot)", "spot",
            "long", round(qty,6), round(ep,4), round(cp,4),
            round(qty*ep,2), round(uPnl,4), 1, 0, 0, 0, "open", ""
        ])

    # options
    for p in pos_options:
        sym  = p["symbol"]
        side = p["side"]
        size = float(p["size"] or 0)
        ep   = float(p.get("avgPrice",0) or 0)
        cp   = float(p.get("markPrice",0) or 0)
        uPnl = float(p.get("unrealisedPnl",0) or 0)
        delta = float(p.get("delta",0) or 0)
        gamma = float(p.get("gamma",0) or 0)
        theta = float(p.get("theta",0) or 0)
        vega  = float(p.get("vega",0)  or 0)
        cid, role = classify(sym, "option", 1 if side=="Buy" else -1)
        leg_id = f"{cid}_{sym[-6:].replace('-','')}"
        legs_rows.append([
            cid, leg_id, role, sym, "option",
            "long" if side=="Buy" else "short",
            size, ep, cp, round(size*ep,4), round(uPnl,4),
            delta, gamma, theta, vega, "open", ""
        ])

    # Пишем строки
    for r_idx, row in enumerate(legs_rows, 2):
        bg = ROW_COLORS[r_idx % 2]
        fmts = [None,None,None,None,None,None,
                "#,##0.0000","#,##0.00","#,##0.00","#,##0.00","#,##0.00",
                "0.0000","0.0000","0.0000","0.0000",None,None]
        style_row(ws_legs, r_idx, bg, row, fmts)

    print(f"  → LEGS: {len(legs_rows)} строк записано")

    # ── Нетто-греки по конструктам ────────────────────────────────────────────
    from collections import defaultdict
    c_agg = defaultdict(lambda: {"uPnl":0,"delta_usd":0,"gamma":0,"theta":0,"vega":0,"notional":0})
    for row in legs_rows:
        cid = row[0]
        if cid in ("?",""):
            continue
        direction = 1 if row[5]=="long" else -1
        qty   = float(row[6] or 0)
        cp    = float(row[8] or 0)
        delta = float(row[11] or 0)
        gamma = float(row[12] or 0)
        theta = float(row[13] or 0)
        vega  = float(row[14] or 0)
        uPnl  = float(row[10] or 0)
        notional = float(row[9] or 0)
        c_agg[cid]["uPnl"]      += uPnl
        c_agg[cid]["delta_usd"] += delta * cp * qty * direction
        c_agg[cid]["gamma"]     += gamma
        c_agg[cid]["theta"]     += theta
        c_agg[cid]["vega"]      += vega
        c_agg[cid]["notional"]  += notional

    # Обновляем CONSTRUCTS
    for r in range(2, ws_con.max_row + 1):
        cid = ws_con.cell(r, 1).value
        if cid in c_agg:
            a = c_agg[cid]
            ws_con.cell(r, 9).value  = round(a["uPnl"],4)
            ws_con.cell(r, 10).value = round(a["delta_usd"],2)
            ws_con.cell(r, 11).value = round(a["gamma"],6)
            ws_con.cell(r, 12).value = round(a["theta"],4)
            ws_con.cell(r, 13).value = round(a["vega"],4)
            ws_con.cell(r, 7).value  = round(a["notional"],2)
    print(f"  → CONSTRUCTS: обновлены греки по {len(c_agg)} конструктам")

    # ── DAILY — снапшот сегодня ───────────────────────────────────────────────
    today = date.today()
    # Найти или добавить строку с сегодняшней датой
    daily_row = None
    for r in range(2, ws_daily.max_row + 2):
        v = ws_daily.cell(r, 1).value
        if v is None:
            daily_row = r; break
        if isinstance(v, (date, datetime)):
            vd = v.date() if isinstance(v, datetime) else v
            if vd == today:
                daily_row = r; break
    if daily_row is None:
        daily_row = ws_daily.max_row + 1

    total_uPnl = sum(float(p.get("unrealisedPnl",0) or 0)
                     for p in pos_linear + pos_options)

    daily_vals = [
        today, round(eq,2), round(im,2), round(mm,2),
        round(total_uPnl,2),
        round(c_agg["C1"]["uPnl"],2),
        round(c_agg["C2"]["uPnl"],2),
        round(c_agg["C3"]["uPnl"],2),
        round(c_agg["C4"]["uPnl"],2),
        0,  # funding_day — нет в снапшоте, заполнить вручную
        round(sum(a["theta"] for a in c_agg.values()),4),
        0,  # basis_C1 — todo
        0,  # z_score_C2 — todo
        0,  # ratio_C3 — todo
        0,  # ratio_z_C3 — todo
        f"auto {datetime.now().strftime('%H:%M')}"
    ]
    fmts_d = ["YYYY-MM-DD"] + ["#,##0.00"]*9 + ["0.0000","0.00","0.00","0.00","0.00",None]
    style_row(ws_daily, daily_row, C_BG, daily_vals, fmts_d)
    print(f"  → DAILY: строка {daily_row} ({today})")

    # ── Сохраняем ─────────────────────────────────────────────────────────────
    wb.save(BOOK_PATH)
    print(f"\n✅  Сохранено: {BOOK_PATH}")

    # ── Итоговая сводка в консоль ─────────────────────────────────────────────
    print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"  Эквити:       {eq:>12,.2f} USDT")
    print(f"  IM:           {im:>12,.2f} USDT")
    print(f"  uPnL total:   {total_uPnl:>12,.4f} USDT")
    print("  P&L по конструктам:")
    for cid in sorted(c_agg):
        print(f"    {cid}: uPnL={c_agg[cid]['uPnl']:>10,.4f}  Δ$={c_agg[cid]['delta_usd']:>10,.2f}  θ={c_agg[cid]['theta']:>8,.4f}")
    unclassified = [r for r in legs_rows if r[0] == "?"]
    if unclassified:
        print(f"\n  ⚠️  Неклассифицировано ({len(unclassified)} позиций):")
        for r in unclassified:
            print(f"      {r[3]:30s} {r[5]:6s} {r[6]}")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

if __name__ == "__main__":
    main()
