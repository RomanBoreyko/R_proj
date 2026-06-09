"""
gen_excel.py — генератор TRADING_BOOK.xlsx
Запуск: python TRADING-SYSTEM/scripts/gen_excel.py
Создаёт файл рядом со скриптом (или по пути OUT_PATH).
"""

import sys
import os
from datetime import date
sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Установи openpyxl: pip install openpyxl")
    sys.exit(1)

# ── пути ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_PATH   = os.path.join(SCRIPT_DIR, "..", "..", "TRADING_BOOK.xlsx")
OUT_PATH   = os.path.normpath(OUT_PATH)

# ── цвета (тёмная тема) ───────────────────────────────────────────────────────
C_BG       = "1E1E2E"   # фон листа (имитация)
C_HEADER   = "2A2A3E"   # шапка
C_ACCENT   = "F0B429"   # золото
C_GREEN    = "3DDC97"
C_RED      = "FF6B6B"
C_BLUE     = "7EB8F7"
C_GRAY     = "3A3A4E"
C_TEXT     = "E0E0F0"
C_WHITE    = "FFFFFF"

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hfont(hex_color=C_TEXT, bold=False, sz=10):
    return Font(color=hex_color, bold=bold, size=sz, name="Consolas")

def halign(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="444466")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, text, width=18):
    cell.value = text
    cell.fill  = hfill(C_HEADER)
    cell.font  = hfont(C_ACCENT, bold=True, sz=10)
    cell.alignment = halign("center")
    cell.border = thin_border()

def style_cell(cell, value=None, color=C_TEXT, bg=C_BG, bold=False,
               align="left", fmt=None):
    if value is not None:
        cell.value = value
    cell.fill  = hfill(bg)
    cell.font  = hfont(color, bold=bold)
    cell.alignment = halign(align)
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt

def set_col_widths(ws, widths: dict):
    """widths = {col_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell

# ── создаём книгу ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # удаляем дефолтный Sheet

# ══════════════════════════════════════════════════════════════════════════════
# 1. DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DASHBOARD")
ws.sheet_properties.tabColor = C_ACCENT

title_row = [
    ("TRADING BOOK — DASHBOARD", C_ACCENT, True, 16),
]

def section_title(ws, row, col, text):
    end_col = col + 5
    # сначала merge, потом пишем в верхнюю левую ячейку
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col)  # верхняя левая — не MergedCell
    c.value = text
    c.fill = hfill(C_GRAY)
    c.font = Font(color=C_ACCENT, bold=True, size=11, name="Consolas")
    c.alignment = halign("left")

# Заголовок
ws.merge_cells("A1:H1")
c = ws["A1"]
c.value = "⚡ TRADING BOOK — DASHBOARD"
c.fill  = hfill(C_HEADER)
c.font  = Font(color=C_ACCENT, bold=True, size=14, name="Consolas")
c.alignment = halign("center", "center")
ws.row_dimensions[1].height = 28

# Блок эквити
labels_equity = [
    ("Эквити (USDT)", "=IFERROR(DAILY!B2,\"\")"),
    ("IM использовано", "=IFERROR(DAILY!C2,\"\")"),
    ("MM", "=IFERROR(DAILY!D2,\"\")"),
    ("P&L сегодня", "=IFERROR(DAILY!E2,\"\")"),
]
# Блок 1: Портфель (col 1-2, rows 3-7)
section_title(ws, 3, 1, "Портфель")
for i, (lbl, frm) in enumerate(labels_equity):
    r = 4 + i
    style_cell(ws.cell(r, 1), lbl, C_BLUE, C_HEADER, bold=True)
    style_cell(ws.cell(r, 2), frm, C_TEXT, C_BG, fmt="#,##0.00")

# Блок 2: P&L по конструктам (col 1-2, rows 9-14)
labels_cpnl = [
    ("C1 Locked Spread",  "=IFERROR(DAILY!F2,\"\")"),
    ("C2 Stat-Arb Pair",  "=IFERROR(DAILY!G2,\"\")"),
    ("C3 Synth Ratio",    "=IFERROR(DAILY!H2,\"\")"),
    ("C4 Opt Calendar",   "=IFERROR(DAILY!I2,\"\")"),
]
section_title(ws, 9, 1, "P&L по конструктам (день)")
for i, (lbl, frm) in enumerate(labels_cpnl):
    r = 10 + i
    style_cell(ws.cell(r, 1), lbl, C_BLUE, C_HEADER, bold=True)
    style_cell(ws.cell(r, 2), frm, C_TEXT, C_BG, fmt="#,##0.00")

# Блок 3: Риск-светофор (col 1-2, rows 15-20)
section_title(ws, 15, 1, "Риск-индикаторы")
risk_labels = [
    ("Z-score C2 (XAU)",    "=IFERROR(DAILY!L2,\"\")"),
    ("Ratio Z-score C3",    "=IFERROR(DAILY!N2,\"\")"),
    ("Basis C1 (спот-перп)","=IFERROR(DAILY!K2,\"\")"),
    ("Margin ratio",        "=IFERROR(DAILY!C2/DAILY!B2,\"\")"),
]
for i, (lbl, frm) in enumerate(risk_labels):
    r = 16 + i
    style_cell(ws.cell(r, 1), lbl, C_BLUE, C_HEADER, bold=True)
    style_cell(ws.cell(r, 2), frm, C_TEXT, C_BG, fmt="0.00")

set_col_widths(ws, {"A": 24, "B": 16, "C": 4, "D": 24, "E": 16})

# ══════════════════════════════════════════════════════════════════════════════
# 2. CONSTRUCTS
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("CONSTRUCTS")
ws.sheet_properties.tabColor = C_BLUE
freeze(ws, "C2")

headers_c = [
    "id", "type", "name", "status", "open_date", "close_date",
    "notional_usd", "realized_pnl", "unrealized_pnl",
    "net_delta_usd", "net_gamma", "net_theta", "net_vega", "notes"
]
widths_c = [6, 18, 22, 10, 12, 12, 14, 14, 14, 14, 12, 12, 12, 30]
for col, (h, w) in enumerate(zip(headers_c, widths_c), 1):
    style_header(ws.cell(1, col), h)
    ws.column_dimensions[get_column_letter(col)].width = w

# Примеры конструктов
constructs = [
    ["C1", "LOCKED_SPREAD",  "XAUTUSDT Spot/Perp",        "active", date.today(), "", 0, 0, 0, 0, 0, 0, 0, "лонг спот + шорт перп XAUTUSDT"],
    ["C2", "STATARB_PAIR",   "XAUUSDT Rev + XAGUSDT Trend","active", date.today(), "", 0, 0, 0, 0, 0, 0, 0, "реверсия XAU / тренд XAG h*≈0.41"],
    ["C3", "SYNTH_RATIO",    "XAUT/XAG Ratio + Options",   "active", date.today(), "", 0, 0, 0, 0, 0, 0, 0, "ratio convergence + optn wrapper"],
    ["C4", "OPT_CALENDAR",   "XAUTUSDT Opt Calendar ATM",  "active", date.today(), "", 0, 0, 0, 0, 0, 0, 0, "near/far synth future ATM"],
]
row_colors = [C_BG, "1A1A2E"]
for r, row in enumerate(constructs, 2):
    bg = row_colors[r % 2]
    for c, val in enumerate(row, 1):
        cell = ws.cell(r, c, value=val)
        cell.fill = hfill(bg)
        cell.font = hfont(C_ACCENT if c == 1 else C_TEXT)
        cell.alignment = halign()
        cell.border = thin_border()
        if c == 5 and isinstance(val, date):
            cell.number_format = "YYYY-MM-DD"

freeze(ws, "C2")

# ══════════════════════════════════════════════════════════════════════════════
# 3. LEGS
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("LEGS")
ws.sheet_properties.tabColor = C_GREEN
headers_l = [
    "construct_id","leg_id","leg_role","instrument","category",
    "direction","qty","entry_price","current_price","notional_usd",
    "unrealized_pnl","delta","gamma","theta","vega","status","notes"
]
widths_l = [12,10,18,28,14,10,10,14,14,14,14,10,10,10,10,10,30]
for col,(h,w) in enumerate(zip(headers_l,widths_l),1):
    style_header(ws.cell(1,col), h)
    ws.column_dimensions[get_column_letter(col)].width = w

# Примеры ног
leg_examples = [
    ["C1","C1_L1","spot_long",   "XAUTUSDT","spot",  "long", 0, 0, 0,"=G2*H2","=(I2-H2)*G2",  1,   0, 0, 0,"open","спот-сетка лонг"],
    ["C1","C1_L2","perp_short",  "XAUTUSDT","perp",  "short",0, 0, 0,"=G3*H3","=(H3-I3)*G3", -1,   0, 0, 0,"open","перп-сетка шорт"],
    ["C2","C2_L1","reversion_perp","XAUUSDT","perp", "long", 0, 0, 0,"=G4*H4","=(I4-H4)*G4",  1,   0, 0, 0,"open","реверсионная"],
    ["C2","C2_L2","trend_perp",  "XAGUSDT", "perp",  "long", 0, 0, 0,"=G5*H5","=(I5-H5)*G5",  1,   0, 0, 0,"open","трендовая XAG"],
    ["C3","C3_L1","ratio_long",  "XAUTUSDT","perp",  "long", 0, 0, 0,"=G6*H6","=(I6-H6)*G6",  1,   0, 0, 0,"open","ratio long leg"],
    ["C3","C3_L2","ratio_short", "XAGUSDT", "perp",  "short",0, 0, 0,"=G7*H7","=(H7-I7)*G7", -1,   0, 0, 0,"open","ratio short leg h*"],
    ["C3","C3_L3","otm_bought",  "XAUT-?-?-C","option","long",0, 0, 0,"=G8*H8","=(I8-H8)*G8",  0,   0, 0, 0,"open","купленный OTM колл"],
    ["C3","C3_L4","otm_sold",    "XAUT-?-?-P","option","short",0,0, 0,"=G9*H9","=(H9-I9)*G9",  0,   0, 0, 0,"open","проданный OTM пут"],
    ["C4","C4_L1","synth_near",  "XAUT-near-ATM","synth_future","long",0,0,0,"=G10*H10","=(I10-H10)*G10",1,0,0,0,"open","near synth: long C + short P"],
    ["C4","C4_L2","synth_far",   "XAUT-far-ATM", "synth_future","short",0,0,0,"=G11*H11","=(H11-I11)*G11",-1,0,0,0,"open","far synth: short C + long P"],
]
for r, row in enumerate(leg_examples, 2):
    bg = row_colors[r % 2]
    for c, val in enumerate(row, 1):
        cell = ws.cell(r, c, value=val)
        cell.fill = hfill(bg)
        cell.font = hfont(C_ACCENT if c <= 3 else C_TEXT)
        cell.alignment = halign()
        cell.border = thin_border()
        if c in (7,8,9,10,11,12,13,14,15):
            cell.number_format = "#,##0.00"
freeze(ws, "D2")

# ══════════════════════════════════════════════════════════════════════════════
# 4. GRID_FILLS
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("GRID_FILLS")
ws.sheet_properties.tabColor = C_GRAY
headers_gf = [
    "date","construct_id","leg_id","instrument",
    "fill_type","direction","qty","price","value_usd","fee_usd",
    "pnl_usd","grid_level","notes"
]
widths_gf = [18,12,10,16,14,10,10,12,14,12,12,12,30]
for col,(h,w) in enumerate(zip(headers_gf,widths_gf),1):
    style_header(ws.cell(1,col), h)
    ws.column_dimensions[get_column_letter(col)].width = w

# Пример
ws.cell(2,1).value = date.today()
ws.cell(2,1).number_format = "YYYY-MM-DD HH:MM"
example_fill = [None,"C1","C1_L1","XAUTUSDT","limit_buy","buy",
                0.01,4245.0,"=G2*H2",0.12,0,1,"пример — заменить на реальные данные"]
for c,val in enumerate(example_fill,1):
    cell = ws.cell(2,c)
    if val is not None: cell.value = val
    cell.fill = hfill(C_BG)
    cell.font = hfont(C_GRAY if val and "пример" in str(val) else C_TEXT)
    cell.alignment = halign()
    cell.border = thin_border()
freeze(ws, "E2")

# ══════════════════════════════════════════════════════════════════════════════
# 5. OPT_EVENTS
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("OPT_EVENTS")
ws.sheet_properties.tabColor = "A855F7"
headers_oe = [
    "date","construct_id","leg_id","event_type",
    "symbol","option_type","strike","expiry",
    "direction","qty","premium_usd","total_usd","fee_usd",
    "delta","gamma","theta","vega","iv_pct","pnl_usd","notes"
]
widths_oe = [18,12,10,16,26,10,10,12,10,8,12,12,10,8,8,8,8,8,12,30]
for col,(h,w) in enumerate(zip(headers_oe,widths_oe),1):
    style_header(ws.cell(1,col), h)
    ws.column_dimensions[get_column_letter(col)].width = w
freeze(ws, "E2")

# ══════════════════════════════════════════════════════════════════════════════
# 6. DAILY
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DAILY")
ws.sheet_properties.tabColor = C_GREEN
headers_d = [
    "date","equity_usd","im_usd","mm_usd",
    "pnl_day","pnl_C1","pnl_C2","pnl_C3","pnl_C4",
    "funding_day","theta_day",
    "basis_C1","z_score_C2","ratio_C3","ratio_z_C3","notes"
]
widths_d = [12,14,12,12,12,12,12,12,12,12,12,12,12,12,12,30]
for col,(h,w) in enumerate(zip(headers_d,widths_d),1):
    style_header(ws.cell(1,col), h)
    ws.column_dimensions[get_column_letter(col)].width = w
freeze(ws, "B2")

# ══════════════════════════════════════════════════════════════════════════════
# 7. PARAMS
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("PARAMS")
ws.sheet_properties.tabColor = C_ACCENT

def params_block(ws, start_row, construct_id, construct_name, color, params):
    """Рисует блок параметров конструкта."""
    # Заголовок блока
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=3)
    c = ws.cell(start_row, 1, f"[{construct_id}] {construct_name}")
    c.fill = hfill(C_GRAY)
    c.font = Font(color=color, bold=True, size=11, name="Consolas")
    c.alignment = halign("left","center")

    for i, (param, desc, val) in enumerate(params):
        r = start_row + 1 + i
        style_cell(ws.cell(r,1), param, C_ACCENT, C_HEADER, bold=True)
        style_cell(ws.cell(r,2), val,   C_TEXT,   C_BG)
        style_cell(ws.cell(r,3), desc,  "888899",  C_BG)
    return start_row + 1 + len(params) + 1

ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 40

params_C1 = [
    ("spot_lower",       "Нижняя граница спот-сетки $",       4000),
    ("spot_upper",       "Верхняя граница спот-сетки $",       4500),
    ("spot_steps",       "Уровней спот-сетки",                   20),
    ("spot_qty_per_step","XAUT на уровень",                    0.01),
    ("perp_lower",       "Нижняя граница перп-сетки $",        4000),
    ("perp_upper",       "Верхняя граница перп-сетки $",       4500),
    ("perp_steps",       "Уровней перп-сетки",                   20),
    ("perp_qty_per_step","Контрактов на уровень",              0.01),
    ("leverage_perp",    "Плечо перп-ноги",                       5),
]
params_C2 = [
    ("xau_mu",           "Среднее XAUUSDT (последние N баров)", 0),
    ("xau_sigma",        "σ XAUUSDT",                           0),
    ("half_life_bars",   "AR(1) полужизнь (бары)",             38),
    ("k_entry",          "Z-порог входа",                     2.0),
    ("k_exit",           "Z-порог выхода",                    0.5),
    ("xau_qty",          "Размер позиции XAU (контракты)",       0),
    ("h_star",           "Хедж-коэф к XAG",                 0.41),
    ("xag_lower",        "Нижняя граница XAG-сетки $",         60),
    ("xag_upper",        "Верхняя граница XAG-сетки $",        75),
]
params_C3 = [
    ("entry_ratio",      "XAUT/XAG при входе",                  0),
    ("target_ratio",     "Цель — μ ratio",                       0),
    ("ratio_mu",         "Скользящее среднее ratio",             0),
    ("ratio_sigma",      "σ ratio",                              0),
    ("h_star",           "Хедж-коэф h*",                     0.41),
    ("xaut_notional",    "Ноционал XAUT $",                      0),
    ("xag_notional",     "Ноционал XAG $ (= xaut × h*)",        0),
    ("otm_call_strikes", "Страйки купленных коллов",         "[]"),
    ("otm_put_strikes",  "Страйки купленных путов",          "[]"),
    ("sold_call_strikes","Страйки проданных коллов",         "[]"),
    ("sold_put_strikes", "Страйки проданных путов",          "[]"),
]
params_C4 = [
    ("near_expiry",      "Ближняя экспирация",       "2026-06-12"),
    ("far_expiry",       "Дальняя экспирация",        "2026-07-31"),
    ("strike_atm",       "ATM страйк (≈ near forward)",      4250),
    ("entry_spread",     "Спред at entry (near-far)",            0),
    ("target_spread",    "Цель при схождении к споту",           0),
    ("qty",              "Лотов",                                 1),
]

row = 1
row = params_block(ws, row, "C1", "Locked Spread — XAUTUSDT Spot/Perp",   C_BLUE,   params_C1)
row = params_block(ws, row, "C2", "Stat-Arb Pair — XAUUSDT Rev + XAG Trend", C_GREEN, params_C2)
row = params_block(ws, row, "C3", "Synth Ratio — XAUT/XAG + Options",     "A855F7", params_C3)
row = params_block(ws, row, "C4", "Opt Calendar — XAUTUSDT ATM",           C_ACCENT, params_C4)

# ══════════════════════════════════════════════════════════════════════════════
# 8. FORMULAS_REF
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("FORMULAS_REF")
ws.sheet_properties.tabColor = C_RED
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 36

formulas = [
    ("ФОРМУЛА",              "ВЫРАЖЕНИЕ",                                     "ПРИМЕЧАНИЕ"),
    ("h* (хедж-коэф)",       "ρ × σ_XAU / σ_XAG",                           "min-variance hedge ratio"),
    ("Z-score",              "(value − μ) / σ",                               "сигнал входа при |Z| > k"),
    ("Basis",                "futures_price − spot_price",                     "+ = contango, − = backwardation"),
    ("Synth future",         "long call(K) + short put(K)",                   "страйк K = ATM forward"),
    ("Calendar spread",      "near_synth_price − far_synth_price",            "собираем при contango"),
    ("$-delta",              "Σ (delta_i × price_i)",                         "НЕ сумма дельт в штуках!"),
    ("Net gamma $",          "Σ (gamma_i × price_i²)",                        "нормировано по ноционалу"),
    ("APR сетки",            "(realized_pnl / investment) / days × 365 × 100","% годовых фактических"),
    ("Locked basis P&L",     "(perp_entry − spot_entry) − (perp_exit − spot_exit)", "схождение спреда"),
    ("Ratio entry signal",   "Z_ratio > +2 → шорт XAUT / лонг XAG",          "ratio уйдёт к μ"),
    ("Half-life (AR1)",      "−ln(2) / ln(β)",                                "β = коэф. AR(1) регрессии"),
    ("IV/RV threshold",      "IV / RV > 1.2 → продаём волу",                 "20% premium к реализованной"),
]

for r, (name, expr, note) in enumerate(formulas, 1):
    bold = (r == 1)
    bg   = C_HEADER if r == 1 else (C_BG if r % 2 == 0 else "1A1A2E")
    fc   = C_ACCENT if r == 1 else C_TEXT
    style_cell(ws.cell(r,1), name, C_ACCENT if r==1 else C_BLUE, bg, bold=bold)
    style_cell(ws.cell(r,2), expr, fc, bg, bold=bold)
    style_cell(ws.cell(r,3), note, "888899", bg)

# ── сохраняем ─────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"✅  Сохранено: {OUT_PATH}")
print(f"    Листы: {', '.join(s.title for s in wb.worksheets)}")
